import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

with open(r'C:\Users\Ryan\.claude\projects\D--CoworkOS-YGO-Project\23bf2266-feb3-40d8-a5eb-3cae23bc6ba0\tool-results\bghvv12q0.txt') as f:
    data = json.load(f)

TARGET = 3
GOLD    = "C9A84C"
DARK_BG = "1E1B2E"
SURF    = "2A2540"
WHITE   = "FFFFFF"

def mk_side(): return Side(style='thin', color="AAAAAA")
def mk_border(): return Border(left=mk_side(), right=mk_side(), top=mk_side(), bottom=mk_side())

wb = Workbook()

# ── Summary sheet ────────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Summary"
ws.sheet_view.showGridLines = False
for col, w in zip("ABCDE", [12, 22, 16, 16, 16]):
    ws.column_dimensions[col].width = w

ws.merge_cells("A1:E1")
ws["A1"] = "RA Series — Prismatic Collector's Rare Restock Report"
ws["A1"].font = Font(name="Arial", bold=True, size=14, color=GOLD)
ws["A1"].fill = PatternFill("solid", fgColor=DARK_BG)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

ws.merge_cells("A2:E2")
ws["A2"] = "Target: 3 copies per card minimum   |   Rarity: Prismatic Collector's Rare"
ws["A2"].font = Font(name="Arial", italic=True, size=10, color="AAAAAA")
ws["A2"].fill = PatternFill("solid", fgColor=DARK_BG)
ws["A2"].alignment = Alignment(horizontal="center")
ws.row_dimensions[2].height = 18

ws.append([])  # row 3 blank

headers = ["Set", "Cards < 3", "Cards = 0", "Copies Needed", "% Complete"]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=4, column=col, value=h)
    c.font = Font(name="Arial", bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=SURF)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = mk_border()
ws.row_dimensions[4].height = 24

sets_data = defaultdict(list)
for r in data:
    sets_data[r["card_number"][:4]].append(r)

data_start = 5
summary_rows = []
for s in sorted(sets_data.keys()):
    cards = sets_data[s]
    under = len(cards)
    zeros = sum(1 for c in cards if c["qty_total"] == 0)
    needed = sum(TARGET - c["qty_total"] for c in cards)
    summary_rows.append((s, under, zeros, needed))

for i, (s, under, zeros, needed) in enumerate(summary_rows):
    row = data_start + i
    fill = "1E2A1E" if i % 2 == 0 else "1A1A2A"
    for col, v in enumerate([s, under, zeros, needed], 1):
        c = ws.cell(row=row, column=col, value=v)
        c.font = Font(name="Arial", color=WHITE)
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(horizontal="center")
        c.border = mk_border()
    pct = ws.cell(row=row, column=5, value=f"=1-(D{row}/(B{row}*3))")
    pct.number_format = "0.0%"
    pct.font = Font(name="Arial", color=WHITE)
    pct.fill = PatternFill("solid", fgColor=fill)
    pct.alignment = Alignment(horizontal="center")
    pct.border = mk_border()

total_row = data_start + len(summary_rows)
ws.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Arial", bold=True, color=GOLD)
ws.cell(row=total_row, column=1).fill = PatternFill("solid", fgColor=SURF)
ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="center")
ws.cell(row=total_row, column=1).border = mk_border()
for col in range(2, 6):
    c = ws.cell(row=total_row, column=col)
    ltr = get_column_letter(col)
    c.value = f"=SUM({ltr}{data_start}:{ltr}{total_row-1})" if col < 5 else f"=1-(D{total_row}/(B{total_row}*3))"
    if col == 5:
        c.number_format = "0.0%"
    c.font = Font(name="Arial", bold=True, color=GOLD)
    c.fill = PatternFill("solid", fgColor=SURF)
    c.alignment = Alignment(horizontal="center")
    c.border = mk_border()

# ── Per-set detail sheets ────────────────────────────────────────────────────
col_widths  = [14, 45, 10, 10, 10, 10, 10, 10, 12, 12]
col_headers = ["Card #", "Card Name", "FE NM", "FE LP", "UN NM", "UN LP",
               "Binder FE", "Binder UN", "Total Qty", "Need to Buy"]

for s in sorted(sets_data.keys()):
    cards = sorted(sets_data[s], key=lambda x: x["card_number"])
    ws2 = wb.create_sheet(s)
    ws2.sheet_view.showGridLines = False
    for i, w in enumerate(col_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    last_col = get_column_letter(len(col_headers))
    ws2.merge_cells(f"A1:{last_col}1")
    ws2["A1"] = f"{s} — Prismatic Collector's Rare Restock List"
    ws2["A1"].font = Font(name="Arial", bold=True, size=13, color=GOLD)
    ws2["A1"].fill = PatternFill("solid", fgColor=DARK_BG)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    for col, h in enumerate(col_headers, 1):
        c = ws2.cell(row=2, column=col, value=h)
        c.font = Font(name="Arial", bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=SURF)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = mk_border()
    ws2.row_dimensions[2].height = 22

    for i, card in enumerate(cards):
        row = 3 + i
        qty = card["qty_total"]
        fill = "2D1010" if qty == 0 else ("2D2010" if qty == 1 else "1A2D1A")
        vals = [
            card["card_number"], card["card_name"],
            card["qty_fe_nm"], card["qty_fe_lp"],
            card["qty_un_nm"], card["qty_un_lp"],
            card["qty_binder_fe_nm"], card["qty_binder_un_nm"],
            qty, TARGET - qty
        ]
        for col, v in enumerate(vals, 1):
            c = ws2.cell(row=row, column=col, value=v)
            c.font = Font(name="Arial", color=WHITE)
            c.fill = PatternFill("solid", fgColor=fill)
            c.border = mk_border()
            if col >= 3:
                c.alignment = Alignment(horizontal="center")

    tr = 3 + len(cards)
    ws2.cell(row=tr, column=1, value="TOTAL").font = Font(name="Arial", bold=True, color=GOLD)
    ws2.cell(row=tr, column=1).fill = PatternFill("solid", fgColor=SURF)
    ws2.cell(row=tr, column=1).border = mk_border()
    ws2.cell(row=tr, column=2, value=f"{len(cards)} cards").font = Font(name="Arial", bold=True, color=GOLD)
    ws2.cell(row=tr, column=2).fill = PatternFill("solid", fgColor=SURF)
    ws2.cell(row=tr, column=2).border = mk_border()
    for col in range(3, len(col_headers) + 1):
        ltr = get_column_letter(col)
        c = ws2.cell(row=tr, column=col, value=f"=SUM({ltr}3:{ltr}{tr-1})")
        c.font = Font(name="Arial", bold=True, color=GOLD)
        c.fill = PatternFill("solid", fgColor=SURF)
        c.alignment = Alignment(horizontal="center")
        c.border = mk_border()

    leg = tr + 2
    ws2.merge_cells(f"A{leg}:D{leg}")
    ws2[f"A{leg}"] = "Legend:"
    ws2[f"A{leg}"].font = Font(name="Arial", bold=True, color="AAAAAA")
    for offset, (label, color) in enumerate([
        ("  Out of stock (qty=0) — buy 3", "2D1010"),
        ("  Critical (qty=1) — buy 2", "2D2010"),
        ("  Almost there (qty=2) — buy 1", "1A2D1A"),
    ]):
        r = leg + 1 + offset
        ws2.merge_cells(f"A{r}:D{r}")
        ws2[f"A{r}"] = label
        ws2[f"A{r}"].font = Font(name="Arial", color=WHITE, size=9)
        ws2[f"A{r}"].fill = PatternFill("solid", fgColor=color)

out_path = r"D:\CoworkOS\YGO Project\scripts\RA-Prismatic-Restock-Report.xlsx"
wb.save(out_path)
print("Saved:", out_path)
